# -*- coding: utf-8 -*-
"""부자재비 오른쪽에 가공비를 나란히 — 받은 엑셀을 고쳐서 저장
   B~F 부자재 / G 공백 / H~M 가공비 · 설정 패널(K열)은 P열로 이동"""
import io, os, shutil, warnings, re
warnings.filterwarnings('ignore')
import openpyxl
from openpyxl.utils import get_column_letter
from copy import copy
from openpyxl.styles import Border, PatternFill

SRC = os.path.join(os.path.expanduser("~"), "Documents", "카카오톡 받은 파일",
                   "서래바이오_제조견적서_20260830_17.xlsx")
OUT = os.path.join(os.path.expanduser("~"), "Desktop",
                   "서래바이오_제조견적서_양식_v2.xlsx")
shutil.copyfile(SRC, OUT)

wb = openpyxl.load_workbook(OUT)
ws = wb['제조견적서']

def unmerge(rng):
    if str(rng) in [str(x) for x in ws.merged_cells.ranges]:
        ws.unmerge_cells(rng)

def cp(src, dst):
    """서식까지 복사"""
    s, d = ws[src], ws[dst]
    d.value = s.value
    d.font, d.border, d.fill = copy(s.font), copy(s.border), copy(s.fill)
    d.alignment, d.number_format = copy(s.alignment), s.number_format

# ── 1) 설정 패널 K열 → P열 ─────────────────────────────────
for r in range(2, 15):
    for col in ('K',):
        src = col + str(r)
        if ws[src].value is not None:
            cp(src, 'R' + str(r))
            ws[src].value = None
ws.column_dimensions['R'].width = 15
ws.column_dimensions['K'].width = 12

# 설정 셀 참조 갱신 (K3 수율 / K5 인상율 / K11 목표단가 / K13 로트총액)
ref = {'$K$3': '$R$3', '$K$5': '$R$5', '$K$11': '$R$11', '$K$13': '$R$13',
       '$K$7': '$R$7', '$K$9': '$R$9'}
for row in ws.iter_rows(min_row=1, max_row=75, min_col=2, max_col=20):
    for c in row:
        if isinstance(c.value, str) and c.value.startswith('='):
            v = c.value
            for a, b in ref.items():
                v = v.replace(a, b)
            c.value = v

# ── 1-2) 패킹 계산 블록(K~O, 19~35행) → T~X 로 대피 (가공비 자리 비우기)
PACK = {'K': 'T', 'L': 'U', 'M': 'V', 'N': 'W', 'O': 'X'}
for r in range(18, 36):
    for sc, dc in PACK.items():
        src, dst = sc + str(r), dc + str(r)
        if not isinstance(ws[src], openpyxl.cell.cell.MergedCell):
            if ws[src].value is not None or ws[src].has_style:
                cp(src, dst)
            ws[src].value = None
            ws[src].border = Border(); ws[src].fill = PatternFill()
# 대피한 칸끼리의 상호 참조 보정
pshift = {}
for r in range(18, 36):
    for sc, dc in PACK.items():
        pshift[sc + str(r)] = dc + str(r)
for r in range(18, 36):
    for dc in PACK.values():
        c = ws[dc + str(r)]
        if isinstance(c.value, str) and c.value.startswith('='):
            c.value = re.sub(r'\$?[A-Z]{1,2}\$?\d{1,3}',
                             lambda mm: pshift.get(mm.group(0).replace('$', ''), mm.group(0)),
                             c.value)
for col, w in [('T', 12), ('U', 3), ('V', 10), ('W', 13), ('X', 13)]:
    ws.column_dimensions[col].width = w

# ── 2) 가공비 블록(B45:I52) → J36:Q43 로 이동 ──────────────
PROC = [(45, 36), (46, 37), (47, 38), (48, 39), (49, 40), (50, 41), (51, 42), (52, 43)]
COLS = {'B': 'J', 'C': 'K', 'D': 'L', 'E': 'M', 'F': 'N', 'G': 'O', 'H': 'P', 'I': 'Q'}

# 옮기기 전 병합 해제
for rng in [str(x) for x in ws.merged_cells.ranges]:
    m = re.match(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', rng)
    if m and 45 <= int(m.group(2)) <= 52:
        ws.unmerge_cells(rng)

# 목적지(I36:P43)에 걸린 병합도 먼저 해제
for rng in [str(x) for x in ws.merged_cells.ranges]:
    mm = re.match(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', rng)
    if mm and 36 <= int(mm.group(2)) <= 43:
        ws.unmerge_cells(rng)

moved = []
for src_r, dst_r in PROC:
    for sc, dc in COLS.items():
        s, d = sc + str(src_r), dc + str(dst_r)
        if ws[s].value is not None or ws[s].has_style:
            cp(s, d)
            moved.append((s, d))
        ws[s].value = None
    ws.row_dimensions[dst_r].height = ws.row_dimensions[src_r].height

# 옮긴 칸 안의 수식 열참조 보정 (E47→K38 처럼 같은 블록 안을 가리키던 것)
shift = {}
for (s, d) in moved:
    shift[re.match(r'([A-Z]+)(\d+)', s).group(0)] = re.match(r'([A-Z]+)(\d+)', d).group(0)
for _, d in moved:
    v = ws[d].value
    if isinstance(v, str) and v.startswith('='):
        def rep(mm):
            key = mm.group(0).replace('$', '')
            return shift.get(key, mm.group(0))
        ws[d].value = re.sub(r'\$?[A-Z]{1,2}\$?\d{1,3}', rep, v)

# 가공비 소계 새 위치 = M43 (기존 G52)
for row in ws.iter_rows(min_row=1, max_row=75, min_col=2, max_col=20):
    for c in row:
        if isinstance(c.value, str) and '$G$52' in c.value:
            c.value = c.value.replace('$G$52', '$O$43')
        elif isinstance(c.value, str) and c.value.strip() == '=G52':
            c.value = '=O43'

# 부자재 '비 고' 열(H,I)은 가공비와 겹치므로 제거
for r in range(37, 44):
    for rng in [str(x) for x in ws.merged_cells.ranges]:
        m3 = re.match(r'H(\d+):I(\d+)', rng)
        if m3 and int(m3.group(1)) == r:
            ws.unmerge_cells(rng)
    for col in ('H', 'I'):
        c = ws[col + str(r)]
        if not isinstance(c, openpyxl.cell.cell.MergedCell):
            c.value = None

# 열 너비 — 오른쪽 블록
for col, w in [('G', 14), ('H', 12), ('I', 2.3), ('J', 16), ('K', 9), ('L', 9), ('M', 7.5), ('N', 12), ('O', 14), ('P', 2), ('Q', 2)]:
    ws.column_dimensions[col].width = w

# ── 3) 빈 자리 정리 · 인쇄영역 ─────────────────────────────
for r in range(44, 54):
    for rng in [str(x) for x in ws.merged_cells.ranges]:
        m2 = re.match(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', rng)
        if m2 and int(m2.group(2)) == r:
            ws.unmerge_cells(rng)
    for col in 'BCDEFGHIJKLMNOPQ':
        c = ws[col + str(r)]
        if not isinstance(c, openpyxl.cell.cell.MergedCell):
            c.value = None
            c.border = Border()
            c.fill = PatternFill()

ws.print_area = "'제조견적서'!$B$1:$O$70"
ws.page_setup.orientation = 'portrait'
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1
ws.sheet_properties.pageSetUpPr.fitToPage = True

wb.save(OUT)
print('saved', OUT)
